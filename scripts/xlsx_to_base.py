#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
数据链路 · 阶段1: xlsx → 基础 CSV

读取 scripts/galgame_ymgal_v2.xlsx 的 galgame_ymgal 表（343 行，权威游戏列表），
输出 scripts/galgame_base.csv（utf-8-sig），列:
  游戏名, 中文名, 别名, 评分人数, tag, rank

其中 中文名/别名 仅作为后续匹配键与兜底；评分人数 用于难度热度算法；
tag/rank 直接进入最终可编辑 CSV（数据库无对应字段）。
其余列（发行年份/品牌/限制级/脚本/原画/音乐/平均分）一律不使用 xlsx 的值。
"""

import argparse
import csv
import os
import sys

from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_DEFAULT = os.path.join(SCRIPT_DIR, "galgame_ymgal_v2.xlsx")
BASE_DEFAULT = os.path.join(SCRIPT_DIR, "galgame_base.csv")

SHEET_NAME = "galgame_ymgal"
OUT_COLUMNS = ["游戏名", "中文名", "别名", "评分人数", "tag", "rank"]


def cell_str(value):
    if value is None:
        return ""
    return str(value).strip()


def main():
    ap = argparse.ArgumentParser(description="xlsx → 基础 CSV（阶段1）")
    ap.add_argument("--xlsx", default=XLSX_DEFAULT, help=f"输入 xlsx 路径（默认: {XLSX_DEFAULT}）")
    ap.add_argument("--base", default=BASE_DEFAULT, help=f"输出基础 CSV 路径（默认: {BASE_DEFAULT}）")
    args = ap.parse_args()

    if not os.path.exists(args.xlsx):
        sys.exit(f"[错误] 找不到输入文件: {args.xlsx}")

    wb = load_workbook(args.xlsx, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"[错误] sheet '{SHEET_NAME}' 不存在，实际: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        sys.exit("[错误] xlsx 为空")

    header = [cell_str(h) for h in rows[0]]
    required = {"游戏名", "中文名", "别名", "评分人数", "tag", "rank"}
    missing = required - set(header)
    if missing:
        sys.exit(f"[错误] xlsx 表头缺少列: {sorted(missing)}；实际表头: {header}")
    col = {name: header.index(name) for name in required}

    out_rows = []
    skip_empty_name = 0
    for raw in rows[1:]:
        game_name = cell_str(raw[col["游戏名"]])
        if not game_name:
            skip_empty_name += 1
            continue

        scorers_raw = raw[col["评分人数"]]
        try:
            scorers = int(float(scorers_raw)) if scorers_raw not in (None, "") else ""
        except (TypeError, ValueError):
            scorers = ""

        out_rows.append({
            "游戏名": game_name,
            "中文名": cell_str(raw[col["中文名"]]),
            "别名": cell_str(raw[col["别名"]]),
            "评分人数": scorers,
            "tag": cell_str(raw[col["tag"]]),
            "rank": cell_str(raw[col["rank"]]),
        })

    with open(args.base, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=OUT_COLUMNS)
        writer.writeheader()
        writer.writerows(out_rows)

    print(f"[完成] 基础 CSV 输出 {len(out_rows)} 行 -> {os.path.abspath(args.base)}")
    print(f"  空游戏名跳过: {skip_empty_name}")
    print(f"  有评分人数:   {sum(1 for r in out_rows if r['评分人数'] != '')} / {len(out_rows)}")
    print(f"  有tag:        {sum(1 for r in out_rows if r['tag'])} / {len(out_rows)}")
    print(f"  有rank:       {sum(1 for r in out_rows if r['rank'])} / {len(out_rows)}")


if __name__ == "__main__":
    main()
