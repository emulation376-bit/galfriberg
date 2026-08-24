#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
根据 galgame_vndb_tags_merged3.xlsx 的白名单，重新生成 galgame_import.csv 的 tag 列。

评分口径（与 VNDB 官方 tag_vn_calc 完全一致，即 VN 页面上展示的 tag 评分）：
    rating = avg(正票) * (正票数 - 负票数) / (总票数)
其中正票 = vote > 0（1..3），负票 = vote < 0，过滤掉 lie=t / ignore=1 /
无 perm_tag 权限用户的票。仅保留 rating >= 阈值（默认 2.0）的 tag。

用法:
  python scripts/apply_vndb_tags_score.py [--csv scripts/galgame_import.csv]
      [--xlsx scripts/galgame_vndb_tags_merged3.xlsx] [--vndb ../VNDB/db]
      [--out scripts/galgame_import.csv] [--min-score 2.0]
"""

import argparse
import csv
import os
import re
import sys
from collections import defaultdict

from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

from extract_vndb_tags import load_title_index, match_vn  # noqa: E402

CSV_DEFAULT = os.path.join(SCRIPT_DIR, "galgame_import.csv")
XLSX_DEFAULT = os.path.join(SCRIPT_DIR, "galgame_vndb_tags_merged3.xlsx")
VNDB_DEFAULT = os.path.normpath(os.path.join(SCRIPT_DIR, "..", "VNDB", "db"))

BASE_HEADERS = [
    "游戏名", "中文名", "别名", "发行年份", "品牌", "限制级",
    "脚本", "原画", "音乐", "声优", "难度", "平均分", "评分人数", "rank",
]


def load_whitelist(xlsx_path):
    """merged3 -> [(中文翻译, 出现次数, frozenset(gids))]，保持 xlsx 行序。"""
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        num = str(row[0]).strip()
        zh = str(row[2]).strip() if row[2] is not None else ""
        count_raw = row[3]
        try:
            count = int(float(str(count_raw).strip().replace(",", ""))) if count_raw not in (None, "") else 0
        except ValueError:
            count = 0
        src = row[4]
        if src:
            gids = frozenset(g.strip() for g in str(src).replace("、", ",").split(",") if g.strip())
        else:
            gids = frozenset([num])
        if zh:
            entries.append((zh, count, gids))
    wb.close()
    # 出现次数降序，同次数保持 xlsx 行序（python sorted 稳定）
    entries.sort(key=lambda e: -e[1])
    return entries


def load_excluded_users(vndb_dir):
    """perm_tag='f' 的用户 id 集合（官方 tag_vn_calc 会排除这些投票）。"""
    excluded = set()
    path = os.path.join(vndb_dir, "users")
    if not os.path.exists(path):
        return excluded
    with open(path, encoding="utf-8") as f:
        for line in f:
            c = line.rstrip("\n").split("\t")
            if len(c) >= 4 and c[3].strip().lower() in ("f", "0", "false"):
                excluded.add(c[0])
    return excluded


def aggregate_ratings(vndb_dir, wanted_vids, excluded_users):
    """
    按官方公式聚合 tags_vn：
      rating = COALESCE(AVG(vote) FILTER (WHERE vote > 0), 3) * SUM(sign(vote)) / COUNT(vote)
    返回 {vid: {tag: rating}}（仅保留 rating > 0 的，最终阈值在外层过滤）。
    """
    acc = defaultdict(lambda: {"pos_sum": 0, "pos": 0, "neg": 0})
    with open(os.path.join(vndb_dir, "tags_vn"), encoding="utf-8") as f:
        for line in f:
            c = line.rstrip("\n").split("\t")
            vid = c[2]
            if vid not in wanted_vids:
                continue
            if c[6].strip().lower() in ("t", "1"):  # ignore
                continue
            if c[7].strip().lower() in ("t", "1"):  # lie
                continue
            if c[3] in excluded_users:
                continue
            try:
                vote = int(c[4])
            except ValueError:
                continue
            if vote == 0:
                continue
            a = acc[(vid, c[1])]
            if vote > 0:
                a["pos_sum"] += vote
                a["pos"] += 1
            else:
                a["neg"] += 1

    ratings = defaultdict(dict)
    for (vid, tag), a in acc.items():
        total = a["pos"] + a["neg"]
        if total == 0:
            continue
        avg_pos = a["pos_sum"] / a["pos"] if a["pos"] else 3.0
        r = avg_pos * (a["pos"] - a["neg"]) / total
        if r > 0:
            ratings[vid][tag] = r
    return ratings


def main():
    ap = argparse.ArgumentParser(description="按 merged3 白名单 + VNDB 官方评分(>=2.0) 重写 CSV 的 tag 列")
    ap.add_argument("--csv", default=CSV_DEFAULT)
    ap.add_argument("--xlsx", default=XLSX_DEFAULT)
    ap.add_argument("--vndb", default=VNDB_DEFAULT)
    ap.add_argument("--out", default=None, help="输出路径，默认覆盖 --csv")
    ap.add_argument("--min-score", type=float, default=2.0)
    args = ap.parse_args()
    out = args.out or args.csv

    # 1) 读白名单
    whitelist = load_whitelist(args.xlsx)
    print(f"[whitelist] merged3 条目: {len(whitelist)}")

    # 2) 读现有 CSV（保留全部基础列）
    with open(args.csv, encoding="utf-8-sig", newline="") as f:
        reader = list(csv.DictReader(f))
    base_headers = (
        [h for h in reader[0].keys() if not h.startswith("tag")]
        if reader
        else list(BASE_HEADERS)
    )
    rows = []
    for r in reader:
        base = {h: r.get(h, "") for h in base_headers}
        rows.append(base)
    print(f"[csv] 行数: {len(rows)}")

    # 3) 匹配 VNDB：优先用 CSV 里的 vndb_id（权威），缺失时才按标题匹配
    exact, norm_index = load_title_index(args.vndb)
    matched = {}
    unmatched = []
    for r in rows:
        name = r["游戏名"].strip()
        cn = (r["中文名"] or "").strip()
        aliases = [a.strip() for a in (r["别名"] or "").split("、") if a.strip()]
        vid = (r.get("vndb_id") or "").strip()
        if not re.fullmatch(r"v\d+", vid):
            vid = match_vn(name, cn, aliases, exact, norm_index) or ""
        if vid:
            matched[name] = vid
        else:
            unmatched.append(name)
    print(f"[match] VNDB 匹配: {len(matched)} / {len(rows)}")
    for n in unmatched:
        print(f"  未匹配: {n}")

    # 4) 聚合官方评分
    excluded_users = load_excluded_users(args.vndb)
    ratings = aggregate_ratings(args.vndb, set(matched.values()), excluded_users)

    # 5) 逐行生成 tag 列（白名单内且评分 > 阈值；按出现次数降序）
    tag_cols = []
    stats = []
    for r in rows:
        vid = matched.get(r["游戏名"].strip())
        tags = []
        if vid:
            vid_ratings = ratings.get(vid, {})
            for zh, _cnt, gids in whitelist:
                if any(vid_ratings.get(g, 0.0) >= args.min_score for g in gids):
                    tags.append(zh)
        stats.append(len(tags))
        tag_cols.append(tags)

    max_tags = max(stats) if stats else 0
    print(f"[result] 有 tag 的游戏: {sum(1 for n in stats if n)} / {len(rows)}")
    print(f"[result] 平均 tag 数: {sum(stats) / len(rows):.2f}   最大 tag 数: {max_tags}")

    # 6) 写回（utf-8-sig + CRLF，基础列原样保留）
    headers = base_headers + [f"tag{i}" for i in range(1, max_tags + 1)]
    with open(out, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, lineterminator="\r\n")
        w.writerow(headers)
        for r, tags in zip(rows, tag_cols):
            w.writerow([r.get(h, "") for h in base_headers] + tags)
    print(f"[done] 已写入: {os.path.abspath(out)}")

    # 7) 与旧版（备份）对比
    old_csv = os.path.join(os.path.dirname(args.csv), "..", ".devlogs", "galgame_import.csv.bak2")
    old_csv = os.path.normpath(old_csv)
    if os.path.exists(old_csv):
        with open(old_csv, encoding="utf-8-sig", newline="") as f:
            old_rows = list(csv.DictReader(f))
        old_vals = set()
        new_vals = set()
        for r in old_rows:
            old_vals.update(v for k, v in r.items() if k.startswith("tag") and v)
        for r, tags in zip(rows, tag_cols):
            new_vals.update(tags)
        print(f"[diff] 旧版唯一 tag 值: {len(old_vals)}  新版唯一 tag 值: {len(new_vals)}")
        print(f"[diff] 保留: {len(old_vals & new_vals)}  新增: {len(new_vals - old_vals)}  消失: {len(old_vals - new_vals)}")
        print("[diff] 消失的 tag（旧版有、新版无）:")
        for v in sorted(old_vals - new_vals):
            print(f"  - {v}")


if __name__ == "__main__":
    main()
