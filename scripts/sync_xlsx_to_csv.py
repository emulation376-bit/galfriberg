#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""以 游戏名.xlsx 为准，重建 galgame_import.csv / galgame_import.merged.csv。

用法:
  python scripts/sync_xlsx_to_csv.py

随后运行 pnpm data:import 将 CSV 写入数据库。
"""
import csv
import os
import re
import sys

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from extract_vndb_tags import load_title_index, match_vn  # noqa: E402

REPO = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))
XLSX = os.path.join(REPO, "scripts", "游戏名.xlsx")
VNDB = os.path.join(REPO, "VNDB", "db")
OUT = [
    os.path.join(REPO, "scripts", "galgame_import.csv"),
    os.path.join(REPO, "scripts", "galgame_import.merged.csv"),
]

# 系列关系：vid -> 是否 ser/seq/preq
series_vids = set()
with open(os.path.join(VNDB, "vn_relations"), encoding="utf-8") as f:
    for line in f:
        c = line.rstrip("\n").split("\t")
        if len(c) >= 3 and c[2] in ("ser", "seq", "preq"):
            series_vids.add(c[0])

wb = load_workbook(XLSX, data_only=True)
ws = wb.active
header = [c.value for c in ws[1]]
col = {h: i for i, h in enumerate(header)}
tag_cols = [i for i, h in enumerate(header) if isinstance(h, str) and re.fullmatch(r"tag\d+", h.strip())]
exact, norm_index = load_title_index(VNDB)

rows = []
for row in ws.iter_rows(min_row=2):
    name = str(row[col["游戏名"]].value or "").strip()
    if not name:
        continue
    vid = str(row[col["vndb_id"]].value or "").strip()
    if not re.fullmatch(r"v\d+", vid):
        cn = str(row[col["中文名"]].value or "").strip() if "中文名" in col else ""
        aliases = [a.strip() for a in str(row[col["别名"]].value or "").replace("、", ",").split(",") if a.strip()]
        vid = match_vn(name, cn, aliases, exact, norm_index) or ""

    def gv(h):
        if h not in col:
            return ""
        v = row[col[h]].value
        return "" if v is None else v

    rows.append({
        "游戏名": name,
        "vndb_id": vid,
        "中文名": gv("中文名"),
        "别名": gv("别名"),
        "发行年份": gv("发行年份"),
        "品牌": gv("品牌"),
        "限制级": gv("限制级"),
        "脚本": gv("脚本"),
        "原画": gv("原画"),
        "音乐": gv("音乐"),
        "声优": gv("声优"),
        "难度": gv("难度"),
        "平均分": gv("平均分"),
        "评分人数": gv("评分人数"),
        "rank": gv("rank"),
        "系列作": gv("系列作") or ("是" if vid in series_vids else "否"),
        "时长": gv("时长"),
        "时长分钟": gv("时长分钟"),
        "_tags": [str(row[i].value).strip() for i in tag_cols if row[i].value is not None and str(row[i].value).strip()],
    })
wb.close()

max_tags = max(len(r["_tags"]) for r in rows)
fieldnames = ["游戏名", "vndb_id", "中文名", "别名", "发行年份", "品牌", "限制级",
              "脚本", "原画", "音乐", "声优", "难度", "平均分", "评分人数", "rank",
              "系列作", "时长", "时长分钟"] + [f"tag{i}" for i in range(1, max_tags + 1)]

for path in OUT:
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, lineterminator="\r\n")
        w.writeheader()
        for r in rows:
            out = {k: ("" if v is None else v) for k, v in r.items() if k != "_tags"}
            for i, t in enumerate(r["_tags"]):
                out[f"tag{i + 1}"] = t
            w.writerow(out)
    print(f"[sync] 已写入 {len(rows)} 行 -> {os.path.abspath(path)}")

print("[sync] 完成。运行 pnpm data:import 将数据写入数据库。")
