#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
直接对已有的 galgame_vndb_tags.xlsx 的“tag统计”sheet 做标签合并精简，输出新 xlsx。

与 merge_vndb_tags.py 不同：本脚本不重新读取 VNDB 转储，而是基于 xlsx 现有行聚合。
- 一个标签可同时属于多组（如 g14 同时进“战斗”与“机甲”）。
- 合并组内同名标签合并计数；未命中任何组的标签保持原样。
- 说明：基于已有“出现次数”列求和，无法在游戏层面去重；
  如需按游戏去重请用 merge_vndb_tags.py。

用法:
  python scripts/merge_vndb_tags_xlsx.py [--in scripts/galgame_vndb_tags.xlsx] [--out scripts/galgame_vndb_tags_merged2.xlsx]
"""

import argparse
import os
from collections import OrderedDict

from openpyxl import load_workbook, Workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
IN_DEFAULT = os.path.join(SCRIPT_DIR, "galgame_vndb_tags.xlsx")
OUT_DEFAULT = os.path.join(SCRIPT_DIR, "galgame_vndb_tags_merged2.xlsx")

# 组名(中文) -> [英文名, 中文翻译, 来源tag列表]
GROUPS = {
    "现代设定": ["Modern Setting", "现代设定", ["g143", "g60", "g221", "g470"]],
    "奇幻": ["Fantasy", "奇幻", ["g2", "g1897", "g1765", "g994", "g3127"]],
    "战斗": ["Combat", "战斗",
             ["g25", "g63", "g66", "g72", "g36", "g44", "g77", "g78", "g79", "g14", "g80"]],
    "未来设定": ["Future Setting", "未来设定", ["g140", "g62", "g244", "g2786"]],
    "虚拟实境/模拟": ["Virtual Reality / Simulation", "虚拟实境/模拟", ["g442", "g2274"]],
    "科幻": ["Science Fiction", "科幻", ["g105", "g106"]],
    "机甲": ["Mecha", "机甲", ["g377", "g14"]],
}

MERGES = {cn: tags for cn, (en, zh, tags) in GROUPS.items()}

GROUP_CAT = {cn: "content" for cn in GROUPS}

CAT_ZH = {"content": "content", "ero": "erotic", "tech": "tech"}


def build_tag_to_groups():
    g2g = {}
    for grp, tags in MERGES.items():
        for t in tags:
            g2g.setdefault(t, []).append(grp)
    return g2g


def main():
    ap = argparse.ArgumentParser(description="对已有的 tag xlsx 做合并精简")
    ap.add_argument("--in", dest="inp", default=IN_DEFAULT, help=f"输入 xlsx（默认: {IN_DEFAULT}）")
    ap.add_argument("--out", default=OUT_DEFAULT, help=f"输出 xlsx（默认: {OUT_DEFAULT}）")
    args = ap.parse_args()

    moved = build_tag_to_groups()

    wb_in = load_workbook(args.inp)
    ws_in = wb_in["tag统计"]
    rows = list(ws_in.iter_rows(values_only=True))
    header = rows[0]
    data = rows[1:]

    # 汇总：key(组名或原tagID) -> dict(编号, 英文名, 中文翻译, 次数, 来源id列表)
    buckets = OrderedDict()

    def put(key, tag_id, en, zh, cnt):
        b = buckets.setdefault(key, {"en": en, "zh": zh, "cnt": 0, "src": []})
        b["cnt"] += cnt if cnt else 0
        b["src"].append(tag_id)

    for r in data:
        tag_id = r[0]
        en = r[1]
        zh = r[4] or ""  # col4 = 中文翻译
        cnt = r[3] if r[3] else 0
        grps = moved.get(tag_id)
        if grps:
            for g in grps:
                g_en, g_zh = GROUPS[g][0], GROUPS[g][1]
                put(g, tag_id, g_en, g_zh, cnt)
        else:
            put(tag_id, tag_id, en, zh, cnt)

    def src_text(b):
        if b["zh"] in MERGES:
            return "、".join(b["src"])
        return ""

    wb_out = Workbook()
    ws = wb_out.active
    ws.title = "tag统计-合并"
    # 输出列：编号, 英文名, 中文翻译, 出现次数, 来源标签
    out_header = ["编号", "英文名", "中文翻译", "出现次数", "来源标签"]
    ws.append(out_header)
    items = sorted(buckets.items(), key=lambda kv: (-kv[1]["cnt"], kv[1]["en"].lower()))
    for key, b in items:
        ws.append([key, b["en"], b["zh"], b["cnt"], src_text(b)])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb_out.save(args.out)
    print(f"[完成] {len(data)} 行 -> {len(buckets)} 行 -> {os.path.abspath(args.out)}")
    for key, b in items:
        print(f"   {b['cnt']:>4}  {b['en']:<26}  {b['zh']:<8}  <- {src_text(b) or '原标签'}")


if __name__ == "__main__":
    main()